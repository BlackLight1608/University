from django.shortcuts import render
import openpyxl
import subprocess
from .models import File
from .models import analyze
from django.core.files.storage import FileSystemStorage
from .forms import UploadFileForm

def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST , request.FILES)
        myfile = request.FILES['file']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name,myfile )
        return render(request , 'upload.html',{'form':form})
    else:
        form = UploadFileForm()
    return render(request, 'upload.html' , {'form':form})


def change(request):
    if request.method == "POST":
        workbk = request.POST['workbk']
        no_sub = int(request.POST['no_sub'])
        new_file= request.POST['new_file']
        print subprocess.check_output("pwd",shell=True)

        wb = openpyxl.load_workbook('/home/darkman/Desktop/mk/k/media/'+workbk)
        sheet = wb.active
        #max row
        mx_r = sheet.max_row
        #max Column
        mx_c = sheet.max_column
        #start and end range for sheet
        start_range = 'C2'
        end_range=chr(ord('C')+(no_sub-1))+str(mx_r)
        for row in sheet[start_range:end_range]:
            for col in row:
                if(col.value=='A'):
                    sheet[col.coordinate]=9
                elif(col.value=='B'):
                    sheet[col.coordinate]=8
                elif(col.value=='C'):
                    sheet[col.coordinate]=7
                elif(col.value=='D'):
                    sheet[col.coordinate]=6
                elif(col.value=='E'):
                    sheet[col.coordinate]=5
                elif(col.value=='S'):
                    sheet[col.coordinate]=10
                else:
                    sheet[col.coordinate]=10.1
        wb.save(filename="/home/darkman/Desktop/mk/k/media/step1.xlsx")
        workbk1 = '/home/darkman/Desktop/mk/k/media/step1.xlsx'
        #the result is stored in this location in sheet
        wb1 = openpyxl.load_workbook(workbk1)
        sheet1 = wb1.active()
        gpa = chr(ord('C')+(no_sub-1))+str(mx_r)
        credit = list(credit)
        #flag
        k = 2
        if((mx_c-2)==no_sub):
            for i in sheet1[start_range:end_range]:
                sum1 = 0
                crdt = 0
                cr_sum = 0
                for col in i:
                    sum1 = sum1 + ((col.value)*int(credit[crdt]))
                    crdt = crdt  + 1
                    cr_sum = cr_sum + int(credit[crdt])
                val = gpa + str(k)
                k = k+1
                res = round(float(sum1)/cr_sum , 2)
                sheet1[val]=res

        wb1.save(filename="/home/darkman/Desktop/mk/k/media/step2.xlsx")
        return render(request,'analyze.html')
    else:
        return render(request,'analyze.html')
